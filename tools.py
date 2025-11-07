import asyncio
import os
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime

from langchain_core.tools import tool
from langchain_mcp_adapters.client import MultiServerMCPClient
import dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import xmind
from xmind.core.topic import TopicElement

zhipu_api_key = dotenv.get_key(".env", "ZHIPU_API_KEY")

@tool
def get_weather(city: str) -> str:
    """Get weather for a given city."""
    return f"{city}，今天是晴天，温度为 25 摄氏度。!"

def get_zhipu_search_mcp_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "url": "https://open.bigmodel.cn/api/mcp/web_search/sse?Authorization="+zhipu_api_key,
                "transport": "sse",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

# 似乎需要魔法
# def get_tavily_search_tools():
#     client = MultiServerMCPClient(
#         {
#             "search": {
#                 "url": "https://mcp.tavily.com/mcp/?tavilyApiKey=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",
#                 "transport": "streamable_http",
#             }
#         }
#     )
#     tools = asyncio.run(client.get_tools())
#     return tools


# https://app.tavily.com/home
# os.environ["TAVILY_API_KEY"] = dotenv.get_key(".env", "TAVILY_API_KEY")
# # # pin install langchain-tavily
# from langchain_tavily import TavilySearch
# toolSearch = TavilySearch(max_results=2)

def get_playwright_mcp_tools():
    client = MultiServerMCPClient(
        {
            "playwright_mcp": {
                "command": "npx",
                "args": ["@playwright/mcp@latest"],
                "transport": "stdio",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_chrome_devtools_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_devtools_mcp": {
                "command": "npx",
                "args": ["chrome-devtools-mcp@latest", "--headless=false", "--isolated=true"],
                "transport": "stdio",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

# 推荐使用 chrome_mcp_tools 进行浏览器操作
def get_chrome_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_mcp": {
                "url": "http://127.0.0.1:12306/mcp",
                "transport": "streamable_http",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_mcp_server_chart_tools():
    client = MultiServerMCPClient(
        {
            "mcp_chart_server": {
                "command": "npx",
                "args": ["-y", "@antv/mcp-server-chart"],
                "transport": "stdio",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_filesystem_tools():
    client = MultiServerMCPClient(
        {
            "mcp_filemanager": {
                "command": "npx",
                "args": ["-y","filemanager-mcp-server"],
                "transport": "stdio"
                }
            }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_xmind_tools():
    client = MultiServerMCPClient(
        {
            "mcp_xmind": {
                "command": "npx",
                "args": ["xmind-generator-mcp"],
                "transport": "stdio",
                "env": {
                    "outputPath": "/Users/darkringsystem/AI/LimbusCompany/files/TestCases",
                    "autoOpenFile": "false"
                }
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

# 启动命令 EXCEL_FILES_PATH=/Users/darkringsystem/AI/LimbusCompany/files/UIAutoCases uvx excel-mcp-server streamable-http
# 优化，使用stdio模式，自动启动
# https://www.modelscope.cn/mcp/servers/codeyijun/excel-mcp-server

# def get_excel_tools():
#     client = MultiServerMCPClient(
#         {
#             "mcp_excel_servers": {
#                 "url": "http://localhost:8017/mcp",
#                 "transport": "streamable_http"
#                  }
#         }
#     )
#     tools = asyncio.run(client.get_tools())
#     return tools

def get_excel_tools():
    client = MultiServerMCPClient(
    {
        "mcp_excel_servers": {
            "command": "uvx",
            "args": ["excel-mcp-server","stdio"],
            "transport": "stdio"
             }
    }
)
    tools = asyncio.run(client.get_tools())
    return tools


@tool
def save_test_cases_to_excel(
    test_cases: List[Dict[str, Any]],
    file_path: str = None,
    sheet_name: str = "测试用例",
    append: bool = False,
    columns: List[str] = None
) -> str:
    """
    将测试用例保存到Excel文件，支持专业格式化。

    此函数可以创建新的Excel文件或向现有文件追加测试用例数据，
    具有自动格式化、列宽调整和完善的错误处理功能。

    参数:
        test_cases: 测试用例列表，每个测试用例是一个字典。
                   示例结构:
                   [
                       {
                           "用例ID": "TC001",
                           "用例标题": "登录功能测试",
                           "前置条件": "用户已注册",
                           "测试步骤": "1. 打开登录页面\n2. 输入用户名和密码\n3. 点击登录按钮",
                           "预期结果": "成功登录并跳转到首页",
                           "优先级": "高",
                           "用例类型": "功能测试"
                       },
                       ...
                   ]
        file_path: Excel文件保存路径，默认为项目根目录下的 "files/UIAutoCases/test_cases.xlsx"
                  可以传入相对路径或绝对路径
        sheet_name: 工作表名称，默认为 "测试用例"
        append: 是否追加到现有文件，True表示追加，False表示创建新文件。默认为False
        columns: 可选的列名列表，用于指定列的顺序。
                如果为None，则自动从测试用例的键中检测列名。
                示例: ["用例ID", "用例标题", "优先级", "预期结果"]

    返回:
        str: 成功消息（包含文件路径和保存的测试用例数量）或错误消息

    异常:
        不会抛出异常，所有错误都会被捕获并作为错误消息返回

    使用示例:
        >>> test_data = [
        ...     {"用例ID": "TC001", "用例标题": "登录测试", "优先级": "高"},
        ...     {"用例ID": "TC002", "用例标题": "登出测试", "优先级": "中"}
        ... ]
        >>> result = save_test_cases_to_excel(test_data, "我的测试.xlsx")
        >>> print(result)
        成功创建！已保存 2 条测试用例到文件: /path/to/我的测试.xlsx
    """
    try:
        # 处理文件路径
        if file_path is None:
            # 使用默认路径
            project_root = Path(__file__).parent
            file_path = project_root / "files" / "TestCases" / "test_cases.xlsx"
        else:
            file_path = Path(file_path)
            # 如果是相对路径且只是文件名，放到默认目录下
            if not file_path.is_absolute() and len(file_path.parts) == 1:
                project_root = Path(__file__).parent
                file_path = project_root / "files" / "TestCases" / file_path
            # 如果是其他相对路径，基于项目根目录解析
            elif not file_path.is_absolute():
                project_root = Path(__file__).parent
                file_path = project_root / file_path

        # 确保目录存在
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 验证输入数据
        if not test_cases:
            return "错误：测试用例列表为空"

        if not isinstance(test_cases, list):
            return "错误：test_cases 必须是字典列表"

        # 确定要使用的列
        if columns is None:
            # 从所有测试用例中自动检测列名
            all_keys = []
            for case in test_cases:
                if not isinstance(case, dict):
                    return f"错误：每个测试用例必须是字典类型，当前发现 {type(case)}"
                for key in case.keys():
                    if key not in all_keys:
                        all_keys.append(key)
            columns = all_keys

        if not columns:
            return "错误：在测试用例中未找到任何列"

        # 定义Excel样式
        # 表头样式：白色粗体文字，蓝色背景
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 单元格样式：普通字体，左对齐，自动换行
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 边框样式：细黑色边框
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 判断是追加模式还是创建新文件
        if append and file_path.exists():
            # 追加模式：加载现有文件
            try:
                wb = load_workbook(str(file_path))
                if sheet_name in wb.sheetnames:
                    # 工作表已存在，追加到末尾
                    ws = wb[sheet_name]
                    start_row = ws.max_row + 1
                else:
                    # 工作表不存在，创建新工作表
                    ws = wb.create_sheet(sheet_name)
                    start_row = 1
            except Exception as e:
                return f"错误：加载现有文件失败: {str(e)}"
        else:
            # 创建新文件
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            start_row = 1

        # 如果是新工作表，写入表头行
        if start_row == 1:
            for col_idx, header in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            start_row = 2  # 数据从第2行开始

        # 写入测试用例数据
        for row_idx, test_case in enumerate(test_cases, start=start_row):
            for col_idx, column_name in enumerate(columns, start=1):
                # 获取字典中的值，如果不存在则为空字符串
                value = test_case.get(column_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 根据内容类型自动调整列宽
        for col_idx, column_name in enumerate(columns, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter

            # 根据列名模式设置默认宽度
            # ID、优先级、类型、状态等短字段：15个字符宽度
            if any(keyword in column_name.lower() for keyword in ['id', 'priority', 'type', 'status', '优先级', '类型', '状态']):
                ws.column_dimensions[column_letter].width = 15
            # 标题、名称、摘要等中等字段：30个字符宽度
            elif any(keyword in column_name.lower() for keyword in ['title', 'name', 'summary', '标题', '名称', '摘要']):
                ws.column_dimensions[column_letter].width = 30
            # 步骤、结果、描述、详情等长字段：50个字符宽度
            elif any(keyword in column_name.lower() for keyword in ['step', 'result', 'description', 'detail', '步骤', '结果', '描述', '详情', '条件']):
                ws.column_dimensions[column_letter].width = 50
            # 其他字段：25个字符宽度
            else:
                ws.column_dimensions[column_letter].width = 25

        # 设置行高以提高可读性
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 30

        # 保存文件
        try:
            wb.save(str(file_path))
        except PermissionError:
            return f"错误：权限被拒绝。文件 '{file_path}' 可能正在其他程序中打开。"
        except Exception as e:
            return f"错误：保存文件失败: {str(e)}"

        # 生成成功消息
        mode_text = "追加" if (append and file_path.exists()) else "创建"
        abs_path = file_path.resolve()
        return f"成功{mode_text}！已保存 {len(test_cases)} 条测试用例到文件: {abs_path}"

    except Exception as e:
        return f"未预期的错误: {str(e)}"


def _fix_xmind_compatibility(xmind_path: Path) -> None:
    """
    修复XMind文件兼容性问题。

    xmind-sdk生成的文件缺少META-INF/manifest.xml，导致XMind 2020+无法打开。
    此函数将manifest.xml添加到xmind文件中。

    参数:
        xmind_path: XMind文件路径
    """
    import zipfile
    import tempfile
    import shutil
    import logging

    logger = logging.getLogger(__name__)

    # manifest.xml内容
    manifest_content = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<manifest xmlns="urn:xmind:xmap:xmlns:manifest:1.0" password-hint="">
    <file-entry full-path="content.xml" media-type="text/xml"/>
    <file-entry full-path="META-INF/" media-type=""/>
    <file-entry full-path="META-INF/manifest.xml" media-type="text/xml"/>
</manifest>'''

    try:
        logger.debug(f"开始修复XMind兼容性: {xmind_path}")

        # 创建临时目录
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            logger.debug(f"临时目录: {temp_path}")

            # 解压xmind文件（xmind文件本质是zip）
            logger.debug("解压XMind文件...")
            with zipfile.ZipFile(xmind_path, 'r') as zip_ref:
                zip_ref.extractall(temp_path)
            logger.debug("✓ 解压完成")

            # 创建META-INF目录
            meta_inf_dir = temp_path / "META-INF"
            meta_inf_dir.mkdir(exist_ok=True)
            logger.debug(f"✓ 创建META-INF目录: {meta_inf_dir}")

            # 写入manifest.xml
            manifest_file = meta_inf_dir / "manifest.xml"
            manifest_file.write_text(manifest_content, encoding='utf-8')
            logger.debug(f"✓ 写入manifest.xml: {manifest_file}")

            # 重新打包为xmind文件
            logger.debug("重新打包XMind文件...")
            file_count = 0
            with zipfile.ZipFile(xmind_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                for file in temp_path.rglob('*'):
                    if file.is_file():
                        arcname = file.relative_to(temp_path)
                        zip_ref.write(file, arcname)
                        file_count += 1
            logger.debug(f"✓ 重新打包完成，共 {file_count} 个文件")

        logger.info(f"✓ XMind兼容性修复成功: {xmind_path}")

    except Exception as e:
        # 修复失败不影响主流程，只记录错误
        logger.error(f"❌ XMind兼容性修复失败: {e}", exc_info=True)
        print(f"警告：XMind兼容性修复失败: {e}")


@tool
def generate_xmind_from_test_cases(
    test_cases: List[Dict[str, Any]],
    requirement_name: str = "测试用例",
    file_path: Optional[str] = None,
    auto_open: bool = False
) -> str:
    """
    根据测试用例生成XMind思维导图文件。

    此函数将测试用例按照层级结构组织成XMind思维导图，便于可视化查看和管理测试用例。

    参数:
        test_cases: 测试用例列表，每个测试用例是一个字典。
                   必需字段:
                   - "用例ID": 测试用例的唯一标识
                   - "测试模块": 测试所属的模块（如：用户登录、购物车等）
                   - "用例标题": 测试用例的标题
                   可选字段:
                   - "前置条件": 执行测试前的准备工作
                   - "操作步骤": 测试的具体步骤
                   - "预期结果": 期望的测试结果
                   - "优先级": 测试用例的优先级（如：P0、P1、P2）
                   - "用例类型": 测试维度（如：正向功能、异常场景、边界值等）

                   示例:
                   [
                       {
                           "用例ID": "TC-FUNC-001",
                           "测试模块": "用户登录",
                           "用例标题": "验证正确的用户名和密码登录",
                           "前置条件": "用户已注册",
                           "操作步骤": "1. 打开登录页面\n2. 输入正确的用户名和密码\n3. 点击登录按钮",
                           "预期结果": "成功登录并跳转到首页",
                           "优先级": "P0",
                           "用例类型": "正向功能"
                       }
                   ]

        requirement_name: 需求名称，将作为思维导图的中心主题。默认为"测试用例"

        file_path: XMind文件保存路径。
                  - 如果为None，则保存到默认路径: files/TestCases/[需求名称]_测试用例_[时间戳].xmind
                  - 可以传入相对路径或绝对路径
                  - 如果只传入文件名，则保存到默认目录下

        auto_open: 是否在生成后自动打开XMind文件。默认为False

    返回:
        str: 成功消息（包含文件路径）或错误消息

    XMind层级结构:
        L0 (中心主题): [需求名称] 测试用例
        L1 (一级分支): [测试模块] (例如：用户登录、购物车)
        L2 (二级分支): [测试维度/用例类型] (例如：正向功能、异常场景、边界值)
        L3 (三级分支): [优先级] [用例标题] (例如：[P0] 验证正确的用户名和密码登录)
        L4 (子主题):
            - 前置条件: [前置条件内容]
            - 操作步骤: [操作步骤内容]
            - 预期结果: [预期结果内容]

    使用示例:
        >>> test_data = [
        ...     {
        ...         "用例ID": "TC-001",
        ...         "测试模块": "用户登录",
        ...         "用例标题": "正确登录",
        ...         "优先级": "P0",
        ...         "用例类型": "正向功能",
        ...         "前置条件": "用户已注册",
        ...         "操作步骤": "1. 输入用户名\\n2. 输入密码\\n3. 点击登录",
        ...         "预期结果": "登录成功"
        ...     }
        ... ]
        >>> result = generate_xmind_from_test_cases(test_data, "登录功能")
        >>> print(result)
        成功生成XMind文件！已保存 1 条测试用例到: /path/to/登录功能_测试用例_20250110_143022.xmind
    """
    import logging

    # 配置日志
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    # 如果没有处理器，添加一个控制台处理器
    if not logger.handlers:
        handler = logging.StreamHandler()
        handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    try:
        logger.info("=" * 60)
        logger.info("开始生成XMind文件")
        logger.info(f"需求名称: {requirement_name}")
        logger.info(f"测试用例数量: {len(test_cases) if isinstance(test_cases, list) else 'N/A'}")
        logger.info(f"文件路径: {file_path}")
        logger.info(f"自动打开: {auto_open}")
        logger.info("=" * 60)

        # 验证输入数据
        logger.debug("步骤1: 验证输入数据")
        if not test_cases:
            logger.error("测试用例列表为空")
            return "错误：测试用例列表为空"

        if not isinstance(test_cases, list):
            logger.error(f"test_cases类型错误: {type(test_cases)}")
            return "错误：test_cases 必须是字典列表"

        logger.info(f"✓ 输入验证通过，共 {len(test_cases)} 条测试用例")

        # 处理文件路径
        logger.debug("步骤2: 处理文件路径")
        if file_path is None:
            # 使用默认路径，包含时间戳
            project_root = Path(__file__).parent
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{requirement_name}_测试用例_{timestamp}.xmind"
            file_path = project_root / "files" / "TestCases" / filename
            logger.info(f"使用默认路径: {file_path}")
        else:
            file_path = Path(file_path)
            # 如果是相对路径且只是文件名，放到默认目录下
            if not file_path.is_absolute() and len(file_path.parts) == 1:
                project_root = Path(__file__).parent
                file_path = project_root / "files" / "TestCases" / file_path
                logger.info(f"相对路径转换为: {file_path}")
            # 如果是其他相对路径，基于项目根目录解析
            elif not file_path.is_absolute():
                project_root = Path(__file__).parent
                file_path = project_root / file_path
                logger.info(f"相对路径解析为: {file_path}")
            else:
                logger.info(f"使用绝对路径: {file_path}")

        # 确保目录存在
        logger.debug("步骤3: 创建目录")
        file_path.parent.mkdir(parents=True, exist_ok=True)
        logger.info(f"✓ 目录已创建: {file_path.parent}")

        # 创建XMind工作簿
        logger.debug("步骤4: 创建XMind工作簿")
        workbook = xmind.load(str(file_path))
        primary_sheet = workbook.getPrimarySheet()
        primary_sheet.setTitle("测试用例")
        logger.info("✓ XMind工作簿已创建")

        # 设置中心主题
        logger.debug("步骤5: 设置中心主题")
        root_topic = primary_sheet.getRootTopic()
        root_topic.setTitle(f"{requirement_name} 测试用例")
        logger.info(f"✓ 中心主题已设置: {requirement_name} 测试用例")

        # 按测试模块分组
        logger.debug("步骤6: 按测试模块分组")
        modules = {}
        for idx, case in enumerate(test_cases):
            if not isinstance(case, dict):
                logger.error(f"测试用例 #{idx} 类型错误: {type(case)}")
                return f"错误：每个测试用例必须是字典类型，当前发现 {type(case)}"

            module = case.get("测试模块", "未分类模块")
            if module not in modules:
                modules[module] = {}

            # 按用例类型（测试维度）分组
            case_type = case.get("用例类型", "功能测试")
            if case_type not in modules[module]:
                modules[module][case_type] = []

            modules[module][case_type].append(case)

        logger.info(f"✓ 测试用例已分组: {len(modules)} 个模块")
        for module_name, types in modules.items():
            total_cases = sum(len(cases) for cases in types.values())
            logger.debug(f"  - {module_name}: {len(types)} 个类型, {total_cases} 条用例")

        # 构建思维导图结构
        logger.debug("步骤7: 构建思维导图结构")
        for module_idx, (module_name, types) in enumerate(modules.items()):
            # L1: 测试模块
            logger.debug(f"  处理模块 {module_idx + 1}/{len(modules)}: {module_name}")
            module_topic = root_topic.addSubTopic()
            module_topic.setTitle(module_name)

            for type_idx, (type_name, cases) in enumerate(types.items()):
                # L2: 测试维度/用例类型
                logger.debug(f"    处理类型 {type_idx + 1}/{len(types)}: {type_name} ({len(cases)} 条用例)")
                type_topic = module_topic.addSubTopic()
                type_topic.setTitle(type_name)

                for case_idx, case in enumerate(cases):
                    # L3: 用例标题（带优先级）
                    priority = case.get("优先级", "")
                    title = case.get("用例标题", "未命名用例")
                    case_title = f"[{priority}] {title}" if priority else title

                    case_topic = type_topic.addSubTopic()
                    case_topic.setTitle(case_title)

                    # 注意：xmind-sdk 不支持 addLabel 方法
                    # 用例ID已经包含在标题中，无需额外添加标签

                    # L4: 添加详细信息作为子主题
                    # 前置条件
                    precondition = case.get("前置条件", "")
                    if precondition:
                        precondition_topic = case_topic.addSubTopic()
                        precondition_topic.setTitle(f"前置条件: {precondition}")

                    # 操作步骤
                    steps = case.get("操作步骤", "")
                    if steps:
                        steps_topic = case_topic.addSubTopic()
                        steps_topic.setTitle(f"操作步骤: {steps}")

                    # 预期结果
                    expected = case.get("预期结果", "")
                    if expected:
                        expected_topic = case_topic.addSubTopic()
                        expected_topic.setTitle(f"预期结果: {expected}")

                    # 将用例ID作为子主题添加（如果存在）
                    case_id = case.get("用例ID", "")
                    if case_id:
                        id_topic = case_topic.addSubTopic()
                        id_topic.setTitle(f"用例ID: {case_id}")

                    # 将其他字段作为子主题添加
                    for key, value in case.items():
                        if key not in ["用例ID", "测试模块", "用例标题", "前置条件",
                                      "操作步骤", "预期结果", "优先级", "用例类型"]:
                            if value:  # 只添加非空值
                                other_topic = case_topic.addSubTopic()
                                other_topic.setTitle(f"{key}: {value}")

        logger.info("✓ 思维导图结构构建完成")

        # 保存XMind文件
        logger.debug("步骤8: 保存XMind文件")
        try:
            logger.info(f"正在保存文件到: {file_path}")
            xmind.save(workbook, str(file_path))
            logger.info("✓ XMind文件已保存")

            # 修复XMind 2020+兼容性问题：添加manifest.xml
            logger.debug("步骤9: 修复XMind 2020+兼容性")
            _fix_xmind_compatibility(file_path)
            logger.info("✓ XMind兼容性修复完成")

        except PermissionError as e:
            logger.error(f"权限错误: {e}")
            return f"错误：权限被拒绝。文件 '{file_path}' 可能正在其他程序中打开。"
        except Exception as e:
            logger.error(f"保存文件失败: {e}", exc_info=True)
            return f"错误：保存文件失败: {str(e)}"

        # 自动打开文件（如果需要）
        if auto_open:
            logger.debug("步骤10: 自动打开文件")
            try:
                import subprocess
                import platform

                system = platform.system()
                if system == "Darwin":  # macOS
                    subprocess.run(["open", str(file_path)])
                    logger.info("✓ 文件已自动打开 (macOS)")
                elif system == "Windows":
                    subprocess.run(["start", str(file_path)], shell=True)
                    logger.info("✓ 文件已自动打开 (Windows)")
                elif system == "Linux":
                    subprocess.run(["xdg-open", str(file_path)])
                    logger.info("✓ 文件已自动打开 (Linux)")
            except Exception as e:
                # 打开失败不影响主流程
                logger.warning(f"自动打开文件失败: {e}")

        # 生成成功消息
        abs_path = file_path.resolve()
        success_msg = f"成功生成XMind文件！已保存 {len(test_cases)} 条测试用例到: {abs_path}"
        logger.info("=" * 60)
        logger.info("✅ XMind文件生成成功！")
        logger.info(f"文件路径: {abs_path}")
        logger.info(f"测试用例数: {len(test_cases)}")
        logger.info("=" * 60)
        return success_msg

    except Exception as e:
        logger.error("=" * 60)
        logger.error("❌ XMind文件生成失败！")
        logger.error(f"错误类型: {type(e).__name__}")
        logger.error(f"错误信息: {str(e)}")
        logger.error("=" * 60)
        logger.exception("详细错误堆栈:")
        return f"未预期的错误: {str(e)}"

